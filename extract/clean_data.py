__author__ = 'cq'

import sys, getopt
import openpyxl
import sqlite3 as db
import datetime


def create_table_sqlite(conn):
    cursor = conn.cursor()
    cursor.executescript('drop table if exists un_women;')
    create_table_query = 'create table un_women (year text,application_id text,project_name text,institution text, project_location_1 text, ' \
                         'project_location_2 text, summary text, project_details text, name_1 text, name_2 text, ' \
                         'project_team text, dob1 text, email_1 text, file_name text, video_link_website text, ' \
                         'date_time text, other_details text, sector text)'
    cursor.execute(create_table_query)
    conn.commit()


def save_to_sqlite(data, conn):
    cursor = conn.cursor()
    qmarks = ', '.join('?' * len(data))
    columns = ', '.join(data.keys())
    query = ('insert into un_women ({}) VALUES ({})'.format(columns, qmarks)).encode('utf-8')

    cursor.execute(query, data.values())
    conn.commit()


def read_file_return_workbook(filename):
    worksheet = openpyxl.load_workbook(filename)
    return worksheet


def process_other_details(ws, row, max_letter):
    current_letter = 'Q'
    other_details = ''
    while current_letter != max_letter:
        temp_details = ws[current_letter + str(row)].value
        if temp_details is not None:
            if isinstance(temp_details, datetime.datetime):
                temp_details = str(temp_details)
            other_details += '\n' + temp_details.encode('utf-8')
        current_letter = chr(ord(current_letter) + 1)

    return other_details


def clean_field(value):
    ret_value = value if value is not None else ''
    if isinstance(ret_value, str):
        ret_value = unicode(ret_value, "utf-8")
    return ret_value


def process_sheet_2011_2014(ws, conn):
    print 'Processing sheet: ' + ws.title
    highest_row = ws.get_highest_row()
    highest_col = ws.get_highest_column()

    for row in range(2, highest_row + 1):
        data_dict = {"year": clean_field(ws['A' + str(row)].value),
                     "application_id": clean_field(ws['B' + str(row)].value),
                     "project_name": clean_field(ws['C' + str(row)].value),
                     "institution": clean_field(ws['D' + str(row)].value),
                     "project_location_1": clean_field(ws['E' + str(row)].value),
                     "project_location_2": clean_field(ws['f' + str(row)].value),
                     "summary": clean_field(ws['G' + str(row)].value),
                     "project_details": clean_field(ws['H' + str(row)].value),
                     "name_1": clean_field(ws['I' + str(row)].value), "name_2": clean_field(ws['J' + str(row)].value),
                     "project_team": clean_field(ws['K' + str(row)].value),
                     "dob1": clean_field(ws['L' + str(row)].value), "email_1": clean_field(ws['M' + str(row)].value),
                     "file_name": clean_field(ws['N' + str(row)].value),
                     "video_link_website": clean_field(ws['O' + str(row)].value),
                     "date_time": clean_field(ws['P' + str(row)].value),
                     "other_details": clean_field(process_other_details(ws, row, openpyxl.utils.get_column_letter(
                         highest_col))), "sector": ''}

        save_to_sqlite(data_dict, conn)


# return file to read and database to save.
def main(argv):
    spreadsheet_filename = ''
    database_filename = ''

    try:
        opts, args = getopt.getopt(argv, "hi:o:", ["ifile=", "ofile="])
    except getopt.GetoptError:
        print 'process_data.py -i <spreadsheet_file> -o <database_name>'
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print 'process_data.py -i <spreadsheet_file2011-2014, spreadsheet_file2015> -o <database_name>'
            sys.exit()
        elif opt in ("-i", "--ifile"):
            spreadsheet_filename = arg
        elif opt in ("-o", "--ofile"):
            database_filename = arg

    return spreadsheet_filename, database_filename


# everything result that is not result_3, result16 to 22 is included into other details.


def process_other_details_2015(ws_data):
    ret_value = ''
    for key, value in ws_data.iteritems():
        if key == 'result_1':
            ret_value += value
    return ret_value


def process_sector_2015(ws_data):
    ret_value = []
    if ws_data["result_16"] == 1:
        ret_value.append('Education')
    if ws_data["result_17"] == 1:
        ret_value.append('Environment \& Sustainability')

    if ws_data["result_18"] == 1:
        ret_value.append('Science \& Technology')

    if ws_data["result_19"] == 1:
        ret_value.append('Health')

    if ws_data["result_20"] == 1:
        ret_value.append('Entrepreneurship \& Business')
    other = ws_data["result_21"]
    if other == 1:
        ret_value.append(ws_data["result_22"])

    return ', '.join(ret_value)


def process_sheet_2015(sheet2015, conn):
    ws_props = sheet2015.sheet_properties
    tab_color = ws_props.tabColor
    if tab_color is None:
        print 'Processing sheet: ' + sheet2015.title
        highest_row = sheet2015.get_highest_row()
        ws_data = {}
        for row in range(2, highest_row + 1):
            ws_data[sheet2015['A' + str(row)].value] = sheet2015['C' + str(row)].value

        data_dict = {"year": 2015, "application_id": ws_data["id"], "project_name": ws_data["general_1"],
                     "institution": ws_data["organization_1"], "project_location_1": ws_data["contact_8"],
                     "project_location_2": ws_data["organization_6"], "summary": ws_data["general_3"],
                     "project_details": ws_data["result_3"], "email_1": ws_data["contact_12"],
                     "name_1": ws_data["contact_1"] + ' ' + ws_data["contact_2"], "name_2": ws_data["contact_4"],
                     "project_team": ws_data["team_3"], "dob1": '', "file_name": ws_data["docsub_5"],
                     "video_link_website": ws_data["docsub_4"], "date_time": '',
                     "other_details": process_other_details_2015(ws_data), "sector": process_sector_2015(ws_data)}

        save_to_sqlite(data_dict, conn)

# running the script:
# python process_data.py -i project_inspire_2011_2014.xlsx,project_inspire_2015.xlsx -o un_women_data.sqlite

if __name__ == '__main__':

    ws_name, db_name = main(sys.argv[1:])

    ws2011, ws2015 = ws_name.split(",")

    conn = db.connect(db_name)
    create_table_sqlite(conn)

    # process 2011-2014 file
    wb2011 = read_file_return_workbook(ws2011)
    wb2011_sheet_names = wb2011.get_sheet_names()
    del wb2011_sheet_names[0]

    for sheet in wb2011_sheet_names:
        process_sheet_2011_2014(wb2011.get_sheet_by_name(sheet), conn)

    # process 2015 file
    wb2015 = read_file_return_workbook(ws2015)
    wb2015_sheet_names = wb2015.get_sheet_names()

    del wb2015_sheet_names[0]
    for sheet_2015 in wb2015_sheet_names:
        process_sheet_2015(wb2015.get_sheet_by_name(sheet_2015), conn)



